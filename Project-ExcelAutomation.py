# This project uses the file: transactions.xlsx

# the "as xl" means it's just an alias so we can make our code shorter when using the library
import openpyxl as xl

# just open two classes
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell = sheet.cell(1,1)

    #loop from the 2nd row (don't include the header) to the last row
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # now create the chart.
    # min_row is telling us which row to start from (i.e. where the data starts)
    # max_row is telling us to go to the end of the data rows
    # min_col and max_col is indicating which data columns to display
    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    #sheet.add_chart(chart, 'e2')   #This line doesn't work in python3. Tells us where to add the corner of the upper left hand corner of the chart

    wb.save(filename)

process_workbook("transactions.xlsx")